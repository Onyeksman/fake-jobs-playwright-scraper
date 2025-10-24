import asyncio
import logging
from typing import List, Dict, Optional
from datetime import datetime
from urllib.parse import urljoin

import pandas as pd
from bs4 import BeautifulSoup
from playwright.async_api import async_playwright, TimeoutError as pwTimeout
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =========================
# CONFIG
# =========================
BASE_URL = "https://realpython.github.io/fake-jobs/"
OUTPUT_FILE = "fake_jobs.xlsx"
NAV_TIMEOUT_MS = 15000
DETAIL_TIMEOUT_MS = 5000
RETRY_ATTEMPTS = 3
LOG_LEVEL = logging.INFO

logging.basicConfig(format="%(asctime)s | %(levelname)s | %(message)s", level=LOG_LEVEL)
logger = logging.getLogger("fake-jobs-scraper")


# =========================
# HELPERS / PARSERS
# =========================
def parse_jobs_from_page(html: str) -> List[Dict]:
    soup = BeautifulSoup(html, "html.parser")
    cards = soup.select("div.card-content")
    jobs = []
    for card in cards:
        title_el = card.select_one("h2.title.is-5")
        company_el = card.select_one("h3.subtitle.is-6.company")
        date_el = card.select_one("p time[datetime]")
        logo_el = card.select_one("figure.image.is-48x48 img")
        link_el = card.find_next("a", string="Apply")

        title = title_el.get_text(strip=True) if title_el else ""
        company = company_el.get_text(strip=True) if company_el else ""
        date_posted = date_el["datetime"].strip() if date_el and date_el.has_attr("datetime") else ""
        logo = logo_el["src"].strip() if logo_el and logo_el.has_attr("src") else ""
        detail_href = link_el["href"].strip() if link_el and link_el.has_attr("href") else ""

        jobs.append({
            "Job Title": title,
            "Company Name": company,
            "Job Detail URL": urljoin(BASE_URL, detail_href) if detail_href else "",
            "Date Posted": date_posted,
            "Logo URL": urljoin(BASE_URL, logo) if logo else "",
        })
    return jobs


def parse_job_details(html: Optional[str]) -> Dict[str, str]:
    """Extract both location and job description from detail page, removing any 'Location: ' prefix."""
    if not html:
        return {"Location": "", "Job Description": ""}
    soup = BeautifulSoup(html, "html.parser")

    location_el = soup.select_one("p#location")
    desc_el = soup.select_one("div.box p")

    location = location_el.get_text(" ", strip=True) if location_el else ""
    if location.lower().startswith("location:"):
        location = location.split(":", 1)[1].strip()

    description = desc_el.get_text(" ", strip=True) if desc_el else ""

    return {"Location": location, "Job Description": description}


# =========================
# DETAIL FETCH (with retries)
# =========================
@retry(stop=stop_after_attempt(RETRY_ATTEMPTS),
       wait=wait_exponential(multiplier=1, min=1, max=5),
       retry=retry_if_exception_type(Exception),
       reraise=True)
async def fetch_detail_html(context, url: str) -> Optional[str]:
    if not url:
        return None
    page = await context.new_page()
    try:
        await page.goto(url, timeout=DETAIL_TIMEOUT_MS, wait_until="domcontentloaded")
        await asyncio.sleep(0.1)
        return await page.content()
    finally:
        try:
            await page.close()
        except Exception:
            pass


# =========================
# MAIN SCRAPER (pagination, preserve exact order)
# =========================
async def main():
    all_jobs: List[Dict] = []

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context()
        page = await context.new_page()

        next_page_url = BASE_URL
        page_index = 1

        while next_page_url:
            logger.info(f"Loading page {page_index}: {next_page_url}")
            try:
                await page.goto(next_page_url, timeout=NAV_TIMEOUT_MS, wait_until="domcontentloaded")
            except pwTimeout:
                logger.warning(f"Timeout loading page {page_index}. Stopping pagination.")
                break
            except Exception as e:
                logger.error(f"Error loading page {page_index}: {e}")
                break

            html = await page.content()
            jobs_on_page = parse_jobs_from_page(html)
            logger.info(f"Found {len(jobs_on_page)} job cards on page {page_index}.")
            all_jobs.extend(jobs_on_page)

            soup = BeautifulSoup(html, "html.parser")
            next_link = (
                soup.select_one("a.pagination-next")
                or soup.select_one("a.next")
                or soup.select_one('a[rel="next"]')
            )
            if next_link and next_link.has_attr("href"):
                next_page_url = urljoin(BASE_URL, next_link["href"])
                page_index += 1
                await asyncio.sleep(0.2)
            else:
                next_page_url = None

        logger.info(f"Total job cards collected (pre-details): {len(all_jobs)}")

        detailed_jobs: List[Dict] = []
        for idx, job in enumerate(all_jobs, start=1):
            detail_url = job.get("Job Detail URL", "") or ""
            details = {"Location": "", "Job Description": ""}
            if detail_url:
                try:
                    html = await fetch_detail_html(context, detail_url)
                    details = parse_job_details(html)
                except pwTimeout:
                    logger.warning(f"Timeout fetching detail for job {idx}: {detail_url}")
                except Exception as e:
                    logger.warning(f"Error fetching detail for job {idx}: {e}")

            cleaned = {
                "Job Title": job.get("Job Title", "").strip(),
                "Company Name": job.get("Company Name", "").strip(),
                "Location": details.get("Location", "").strip(),
                "Date Posted": job.get("Date Posted", "").strip(),
                "Logo URL": job.get("Logo URL", "").strip(),
                "Job Detail URL": detail_url.strip(),
                "Job Description": details.get("Job Description", "").strip(),
            }
            detailed_jobs.append(cleaned)

            if idx % 10 == 0 or idx == len(all_jobs):
                logger.info(f"Processed {idx}/{len(all_jobs)} job details...")

        await context.close()
        await browser.close()

    # =========================
    # DATAFRAME CLEANING
    # =========================
    df = pd.DataFrame(detailed_jobs, columns=[
        "Job Title", "Company Name", "Location", "Date Posted",
        "Logo URL", "Job Detail URL", "Job Description"
    ])

    df = df.applymap(lambda v: v.strip() if isinstance(v, str) else v)
    df.replace(["", None], "N/A", inplace=True)
    df.drop_duplicates(keep="first", inplace=True)
    df.reset_index(drop=True, inplace=True)

    df.to_excel(OUTPUT_FILE, index=False)
    style_excel(OUTPUT_FILE)
    logger.info(f"✅ Saved {len(df)} job records to {OUTPUT_FILE}")


# =========================
# EXCEL STYLING
# =========================
def style_excel(file_path: str):
    wb = load_workbook(file_path)
    ws = wb.active

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    grey_font = Font(color="808080", italic=True)
    alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    border_side = Side(border_style="medium", color="000000")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    # Header
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if r_idx % 2 == 0:
            for cell in row:
                cell.fill = alt_fill
        for cell in row:
            try:
                if str(cell.value).strip() == "N/A":
                    cell.font = grey_font
            except Exception:
                pass
            cell.border = border

    # Auto-fit columns
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                v = str(cell.value) if cell.value is not None else ""
                if len(v) > max_length:
                    max_length = len(v)
            except Exception:
                pass
        adjusted_width = min((max_length + 2), 80)
        ws.column_dimensions[col_letter].width = adjusted_width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    last_row = ws.max_row + 2
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    src_cell = ws.cell(row=last_row, column=1, value="📊 Sourced from (https://realpython.github.io/fake-jobs/)")
    time_cell = ws.cell(row=last_row + 1, column=1, value=f"Generated on: {timestamp}")
    src_cell.font = grey_font
    time_cell.font = grey_font

    wb.save(file_path)
    wb.close()


# =========================
# ENTRY POINT
# =========================
if __name__ == "__main__":
    asyncio.run(main())
